"""
export_review_excel.py
----------------------
Convert cluster_candidates_review_filtered.csv to formatted Excel.
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INPUT  = Path("outputs/cluster_candidates_review_filtered.csv")
OUTPUT = Path("outputs/cluster_review.xlsx")

COL_WIDTHS = {
    "term":           40,
    "document_count": 12,
    "percent":        10,
    "status":         10,
    "ziel_cluster":   25,
    "kommentar":      40,
}

FILL_RED   = PatternFill("solid", fgColor="FFB3B3")
FILL_GREEN = PatternFill("solid", fgColor="B3FFB3")
FILL_BLUE  = PatternFill("solid", fgColor="B3D9FF")
FILL_NOISE = PatternFill("solid", fgColor="E0E0E0")


def main():
    df = pd.read_csv(INPUT, encoding="utf-8")
    df["status"]       = df["status"].fillna("")
    df["ziel_cluster"] = df["ziel_cluster"].fillna("")
    df["kommentar"]    = df["kommentar"].fillna("")

    df.to_excel(OUTPUT, index=False, engine="openpyxl")
    wb = load_workbook(OUTPUT)
    ws = wb.active

    # Bold header + background
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze first row
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        width = COL_WIDTHS.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation on "status" column (col D = index 4)
    status_col = df.columns.get_loc("status") + 1
    dv = DataValidation(
        type="list",
        formula1='"0,1,2"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(df)+1}"
    ws.add_data_validation(dv)

    # Row color coding based on status value
    for row_idx in range(2, len(df) + 2):
        status_cell = ws.cell(row=row_idx, column=status_col)
        val = str(status_cell.value).strip()

        if val == "0":
            fill = FILL_RED
        elif val == "1":
            fill = FILL_GREEN
        elif val == "2":
            fill = FILL_BLUE
        elif val == "NOISE_SUSPECT":
            fill = FILL_NOISE
        else:
            continue

        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Row height
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f"Saved {len(df)} rows to {OUTPUT}")


if __name__ == "__main__":
    main()
