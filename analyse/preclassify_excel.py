"""
preclassify_excel.py
--------------------
Auto-classify terms in cluster_review.xlsx.
status = 0 → noise
status = 2 → existing cluster
status = "" → potential new cluster (manual review needed)
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT  = Path("outputs/cluster_review.xlsx")
OUTPUT = Path("outputs/cluster_review_preclassified.xlsx")

# ---------------------------------------------------------------------------
# Noise patterns
# ---------------------------------------------------------------------------

NOISE_WORDS = {
    "hätte", "wäre", "wurde", "wurden", "seien", "geworden",
    "berichtet", "beschreibt", "handelt", "stellt", "stellen",
    "kurzgefasst", "gehabt", "bereits", "winter", "sommer",
    "herbst", "frühling", "vorliegenden", "unterlagen",
    "zusammenfassend", "hinaus", "erhoffe", "erfahre",
    "betrage", "bestehe", "würde", "würden", "welche",
    "zusammenhang", "hinsichtlich", "können", "genug",
    "zurückhaltend", "schnell", "meistens", "allem", "allen",
    "beeinträchtigt", "gestört", "treten", "beschreibt",
}

# ---------------------------------------------------------------------------
# Known cluster keyword map
# ---------------------------------------------------------------------------

KNOWN_CLUSTERS = {
    "LWS_SYNDROM": [
        "lumboischialgie", "lumbago", "lumbalgie", "lumbalgien",
        "radikulär", "ischialgie", "ischias",
    ],
    "WS_SYNDROM": [
        "dorsalgie", "dorsalgien", "rückenschmerz", "wirbelsäule",
        "spondylose", "spondylarthrose", "spondylitis",
        "osteochondrose", "osteochondrosen", "ostechondrose",
        "hyperkyphose", "kyphose", "skoliose", "lordose",
        "bandscheibe", "diskushernie",
    ],
    "HWS_SYNDROM": [
        "zervikobrachialgien", "zervikozephale", "zervikalsyndrom",
        "zervikogen",
    ],
    "TINNITUS_AURIUM": [
        "tinnitus", "pfeifen", "rauschen",
    ],
    "MIGRAENE_KOPFSCHMERZEN": [
        "kopfschmerz", "migräne", "cephalgie", "cephalgien",
        "spannungskopfschmerz", "migräneattacken", "kopfschmerzarten",
    ],
    "REIZDARM": [
        "reizdarm", "durchfall", "obstipation", "bauchschmerz",
        "colitis", "zöliakie", "crohn", "gastritis",
    ],
    "MUEDIGKEIT_ERSCHOEPFUNG": [
        "erschöpfung", "erschöpfungs", "müdigkeit", "fatigue",
        "burnout", "erschöpfungssyndrom", "erschöpfungsgefühl",
        "erschöpfungssymptom",
    ],
    "SCHLAFSTOERUNGEN": [
        "schlafstörung", "insomnia", "durchschlaf", "einschlaf",
        "schlafapnoe",
    ],
    "FIBROMYALGIE": ["fibromyalgie"],
    "POLYNEUROPATHIE": [
        "polyneuropathie", "polyneuropath", "neuropathie",
        "neuropathisch", "karpaltunnel",
    ],
    "GELENKE": [
        "arthrose", "arthritis", "koxarthrose", "gonarthrose",
        "rhizarthrose", "fingerpolyarthrose", "polyarthrose",
        "polyarthritis", "rheuma", "rheumatica", "polymyalgia",
        "impingement", "bursitis", "tendinitis",
    ],
    "PSYCHE_KOGNITION": [
        "depression", "angststörung", "gedächtnisstörung",
        "konzentrationsstörung", "merkfähigkeitsstörung",
        "konzentrationsst",
    ],
    "SCHILDDRÜSE": [
        "thyreoiditis", "hashimoto", "basedow",
        "hypothyreose", "hyperthyreose",
    ],
    "HAUT": [
        "neurodermitis", "dermatitis", "psoriasis", "ekzem", "urtikaria",
    ],
    "HNO": ["rhinitis", "sinusitis", "pharyngitis"],
    "HERZ_KREISLAUF": [
        "hypertonie", "hypotonie", "arrhythmie", "herzinsuffizienz",
        "tachykardie", "bradykardie",
    ],
    "DIABETES": ["diabetes"],
    "RESTLESS_LEGS": ["restless", "legs syndrom"],
    "RAYNAUD": ["raynaud"],
    "OSTEOPOROSE": ["osteoporose", "osteopenie"],
    "ATEMWEGE": ["bronchitis", "asthma"],
    "SEHSTOERUNGEN": ["sehstörung"],
}

# Fills
FILL_RED   = PatternFill("solid", fgColor="FFB3B3")
FILL_BLUE  = PatternFill("solid", fgColor="B3D9FF")
FILL_GREY  = PatternFill("solid", fgColor="E0E0E0")
FILL_NONE  = PatternFill(fill_type=None)


def classify(term: str, current: str) -> str:
    """Return new status or keep existing."""
    c = str(current).strip()
    # Don't overwrite 0, 1, 2
    if c in ("0", "1", "2"):
        return c

    t = term.lower()
    parts = t.split()

    # Noise: NOISE_SUSPECT flag
    if c == "NOISE_SUSPECT":
        return "0"

    # Noise: too many words
    if len(parts) > 3:
        return "0"

    # Noise: too long
    if len(term) > 40:
        return "0"

    # Noise: contains noise words
    if any(p in NOISE_WORDS for p in parts):
        return "0"

    # Existing cluster match
    for keywords in KNOWN_CLUSTERS.values():
        if any(kw in t for kw in keywords):
            return "2"

    return ""


def apply_fill(ws, row_idx: int, ncols: int, status: str):
    if status == "0":
        fill = FILL_GREY
    elif status == "2":
        fill = FILL_BLUE
    else:
        fill = FILL_NONE
    for col in range(1, ncols + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def main():
    # Load with pandas to manipulate data
    df = pd.read_excel(INPUT, dtype=object)
    df = df.fillna("").astype(str)
    # dtype=str converts NaN to "nan" string — clean that up
    df = df.replace("nan", "")

    # Apply classification
    df["status"] = df.apply(
        lambda r: classify(r["term"], r["status"]), axis=1
    )

    # Save updated data back to Excel
    df.to_excel(OUTPUT, index=False, engine="openpyxl")

    # Re-open with openpyxl for formatting
    wb = load_workbook(OUTPUT)
    ws = wb.active
    ncols = len(df.columns)

    # Bold header
    from openpyxl.styles import Font, Alignment
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    col_widths = {
        "term": 40, "document_count": 12, "percent": 10,
        "status": 10, "ziel_cluster": 25, "kommentar": 40,
    }
    from openpyxl.utils import get_column_letter
    for idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col_name, 20)

    # Color rows by status
    for row_idx in range(2, len(df) + 2):
        status_val = str(ws.cell(row=row_idx, column=4).value or "").strip()
        apply_fill(ws, row_idx, ncols, status_val)

    wb.save(OUTPUT)

    # Summary
    total        = len(df)
    status_0     = (df["status"] == "0").sum()
    status_2     = (df["status"] == "2").sum()
    unclassified = (df["status"] == "").sum()

    print(f"Total:               {total}")
    print(f"Status 0 (noise):    {status_0}")
    print(f"Status 2 (existing): {status_2}")
    print(f"Unclassified (new):  {unclassified}")
    print(f"\nSaved to {OUTPUT}")

    print("\nPotential NEW cluster candidates:")
    print("-" * 45)
    new = df[df["status"] == ""].sort_values("document_count", ascending=False)
    print(new[["term", "document_count", "percent"]].to_string(index=False))


if __name__ == "__main__":
    main()
